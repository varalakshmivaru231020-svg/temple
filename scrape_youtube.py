import yt_dlp
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

CHANNEL_URL = "https://www.youtube.com/@SriSurabharathiTV/videos"
OUTPUT_FILE = r"c:\Users\varalakshmi\Desktop\Surabharathi Temple\SriSurabharathiTV_Videos.xlsx"

print("Fetching video list from YouTube channel... (this may take a minute)")

ydl_opts = {
    "quiet": True,
    "extract_flat": True,
    "skip_download": True,
    "ignoreerrors": True,
}

videos = []
with yt_dlp.YoutubeDL(ydl_opts) as ydl:
    info = ydl.extract_info(CHANNEL_URL, download=False)
    if info and "entries" in info:
        entries = list(info["entries"])
        print(f"Found {len(entries)} videos. Fetching details...")
        for i, entry in enumerate(entries, 1):
            if not entry:
                continue
            video_id = entry.get("id", "")
            title = entry.get("title", "N/A")
            url = f"https://www.youtube.com/watch?v={video_id}" if video_id else entry.get("url", "N/A")
            category = entry.get("categories", [None])[0] if entry.get("categories") else "N/A"
            upload_date = entry.get("upload_date", "")
            if upload_date and len(upload_date) == 8:
                upload_date = f"{upload_date[:4]}-{upload_date[4:6]}-{upload_date[6:]}"
            duration = entry.get("duration", 0)
            if duration:
                mins, secs = divmod(int(duration), 60)
                hrs, mins = divmod(mins, 60)
                duration_str = f"{hrs}:{mins:02d}:{secs:02d}" if hrs else f"{mins}:{secs:02d}"
            else:
                duration_str = "N/A"
            videos.append({
                "no": i,
                "title": title,
                "url": url,
                "category": category,
                "upload_date": upload_date or "N/A",
                "duration": duration_str,
            })
            print(f"  [{i}/{len(entries)}] {title[:60]}")

# If categories are all N/A, try to fetch individual video details for first 20
if all(v["category"] == "N/A" for v in videos) and videos:
    print("\nCategories not in flat extract — fetching individual video details for categories...")
    detail_opts = {"quiet": True, "skip_download": True, "ignoreerrors": True}
    with yt_dlp.YoutubeDL(detail_opts) as ydl:
        for v in videos[:50]:
            try:
                detail = ydl.extract_info(v["url"], download=False)
                if detail:
                    cats = detail.get("categories", [])
                    v["category"] = cats[0] if cats else "N/A"
                    if not v["upload_date"] or v["upload_date"] == "N/A":
                        ud = detail.get("upload_date", "")
                        if ud and len(ud) == 8:
                            v["upload_date"] = f"{ud[:4]}-{ud[4:6]}-{ud[6:]}"
            except Exception:
                pass

# Build Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Videos"

header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
headers = ["#", "Video Title", "Video Link", "Category", "Upload Date", "Duration"]
col_widths = [5, 60, 55, 25, 15, 12]

for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions[cell.column_letter].width = w

ws.row_dimensions[1].height = 22

alt_fill = PatternFill(start_color="EAF0FB", end_color="EAF0FB", fill_type="solid")
link_font = Font(color="1155CC", underline="single")

for row_idx, v in enumerate(videos, 2):
    fill = alt_fill if row_idx % 2 == 0 else None
    ws.cell(row=row_idx, column=1, value=v["no"])
    ws.cell(row=row_idx, column=2, value=v["title"])
    link_cell = ws.cell(row=row_idx, column=3, value=v["url"])
    link_cell.hyperlink = v["url"]
    link_cell.font = link_font
    ws.cell(row=row_idx, column=4, value=v["category"])
    ws.cell(row=row_idx, column=5, value=v["upload_date"])
    ws.cell(row=row_idx, column=6, value=v["duration"])
    if fill:
        for col in range(1, 7):
            ws.cell(row=row_idx, column=col).fill = fill

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:F{len(videos)+1}"

wb.save(OUTPUT_FILE)
print(f"\nDone! {len(videos)} videos saved to:\n{OUTPUT_FILE}")
