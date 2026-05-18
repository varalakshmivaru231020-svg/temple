"""
Reads SriSurabharathiTV_Videos.xlsx, downloads YouTube thumbnails,
converts them to WebP, saves to the backend uploads folder,
and inserts all videos into MongoDB VideoGalleryItem collection.
"""

import openpyxl
import requests
import time
import os
import re
from io import BytesIO
from PIL import Image
from pymongo import MongoClient
from datetime import datetime, timezone

# ── CONFIG ──────────────────────────────────────────────────────────────
EXCEL_FILE   = r"c:\Users\varalakshmi\Desktop\Surabharathi Temple\SriSurabharathiTV_Videos.xlsx"
UPLOADS_DIR  = r"c:\Users\varalakshmi\Desktop\Surabharathi Temple\temple_backend\uploads"
MONGO_URI    = "mongodb+srv://rasvikas124verma_db_user:VPMolqJhzu5NO2Mt@cluster0.sqg2nyn.mongodb.net/temple_db?appName=Cluster0"
DB_NAME      = "temple_db"
COLLECTION   = "videogalleryitems"
# ────────────────────────────────────────────────────────────────────────

os.makedirs(UPLOADS_DIR, exist_ok=True)

def extract_video_id(url):
    m = re.search(r"(?:v=|youtu\.be/)([A-Za-z0-9_-]{11})", url or "")
    return m.group(1) if m else None

def download_thumbnail_webp(video_id, filename):
    """Try maxresdefault → hqdefault → mqdefault. Returns saved filepath or None."""
    for quality in ("maxresdefault", "hqdefault", "mqdefault"):
        thumb_url = f"https://img.youtube.com/vi/{video_id}/{quality}.jpg"
        try:
            r = requests.get(thumb_url, timeout=15)
            if r.status_code == 200 and len(r.content) > 5000:
                img = Image.open(BytesIO(r.content)).convert("RGB")
                out_path = os.path.join(UPLOADS_DIR, filename)
                img.save(out_path, "WEBP", quality=82)
                return filename
        except Exception:
            pass
    return None

# ── Read Excel ───────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb.active

rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    no, title, url, category, upload_date, duration = (list(row) + [None]*6)[:6]
    if not title or not url:
        continue
    rows.append({
        "no": no,
        "title": str(title).strip(),
        "url": str(url).strip(),
        "category": str(category).strip() if category and str(category) != "N/A" else "General",
        "upload_date": str(upload_date).strip() if upload_date else "",
    })

print(f"Read {len(rows)} videos from Excel.\n")

# ── Connect MongoDB ──────────────────────────────────────────────────────
print("Connecting to MongoDB...")
client = MongoClient(MONGO_URI)
db     = client[DB_NAME]
col    = db[COLLECTION]

# Build set of existing URLs to avoid duplicates
existing_urls = set(doc["videoUrl"] for doc in col.find({}, {"videoUrl": 1}))
print(f"Existing videos in DB: {len(existing_urls)}\n")

# ── Process & Insert ─────────────────────────────────────────────────────
inserted = 0
skipped  = 0
no_thumb = 0

for i, v in enumerate(rows, 1):
    if v["url"] in existing_urls:
        print(f"  [{i}/{len(rows)}] SKIP (already exists): {v['title'][:55]}")
        skipped += 1
        continue

    video_id = extract_video_id(v["url"])
    thumbnail_filename = None

    if video_id:
        ts = int(time.time() * 1000) + i          # unique timestamp per video
        safe_name = re.sub(r"[^a-z0-9]+", "-", v["title"].lower())[:40]
        webp_name = f"{ts}-{safe_name}.webp"
        result = download_thumbnail_webp(video_id, webp_name)
        if result:
            thumbnail_filename = result
            print(f"  [{i}/{len(rows)}] thumbnail OK  : {v['title'][:55]}")
        else:
            no_thumb += 1
            print(f"  [{i}/{len(rows)}] no thumbnail  : {v['title'][:55]}")
    else:
        print(f"  [{i}/{len(rows)}] no video_id   : {v['title'][:55]}")

    doc = {
        "title":     v["title"],
        "category":  v["category"],
        "videoUrl":  v["url"],
        "when":      v["upload_date"],
        "thumbnail": thumbnail_filename or "",
        "isPinned":  False,
        "isActive":  True,
        "createdAt": datetime.now(timezone.utc),
        "updatedAt": datetime.now(timezone.utc),
    }
    col.insert_one(doc)
    inserted += 1

# ── Update Excel with thumbnail column ───────────────────────────────────
# Add/update Thumbnail column (column 7)
ws.cell(row=1, column=7, value="Thumbnail File")
from openpyxl.styles import Font, PatternFill, Alignment
ws.cell(row=1, column=7).font = Font(bold=True, color="FFFFFF", size=11)
ws.cell(row=1, column=7).fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
ws.cell(row=1, column=7).alignment = Alignment(horizontal="center")
ws.column_dimensions["G"].width = 45

for idx, v in enumerate(rows, 2):
    video_id = extract_video_id(v["url"])
    if video_id:
        ts = int(time.time() * 1000) + idx
        safe_name = re.sub(r"[^a-z0-9]+", "-", v["title"].lower())[:40]
        # find the actual file saved
        prefix = re.sub(r"[^a-z0-9]+", "-", v["title"].lower())[:40]
        matches = [f for f in os.listdir(UPLOADS_DIR) if prefix in f and f.endswith(".webp")]
        ws.cell(row=idx, column=7, value=matches[0] if matches else "")

wb.save(EXCEL_FILE)

print(f"\n{'='*55}")
print(f"  Done!")
print(f"  Inserted : {inserted} videos")
print(f"  Skipped  : {skipped} (already in DB)")
print(f"  No thumb : {no_thumb}")
print(f"  Excel updated with thumbnail column.")
print(f"{'='*55}")
client.close()
